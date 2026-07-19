"""Spreadsheet chat skill tests: end-to-end over the HTTP API (upload -> chat
with attachment -> kind-specific actions), driven by the demo connector and the
deterministic mock provider like the orchestrator tests.

Each test creates its own project so context merges never leak into other tests.
"""

from __future__ import annotations

import io
import json
import re

from openpyxl import Workbook

from app.agent.intent import detect_intent
from app.agent.skills import SKILLS
from app.db.models import Attachment
from app.services import context_store

# --- helpers -----------------------------------------------------------------


def _xlsx_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coa_workbook() -> Workbook:
    """Members chosen against the MCWPCF demo fixtures: 'Total Payroll',
    'Salaries', 'Overtime' exist in Account; 'wages' differs only by case;
    'Rocket Fuel' does not exist. 'Overtime' is given the wrong parent."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    for row in (
        ["Account", "Parent", "Alias"],
        ["Total Payroll", "", "Payroll Costs"],
        ["Salaries", "Total Payroll", ""],
        ["wages", "Total Payroll", ""],
        ["Overtime", "Total Revenue", ""],
        ["Rocket Fuel", "Total Payroll", ""],
    ):
        ws.append(row)
    return wb


def _layout_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "PayrollLayout"
    ws.append(["Scenario: Actual"])
    ws.append([None, "Jan", "Feb", "Mar"])
    ws.append(["Total Payroll", 100, 110, 120])
    ws.append(["Salaries", 50, 55, 60])
    ws.append(["Wages", 30, 33, 36])
    ws.append(["Rocket Fuel", 1, 2, 3])
    return wb


def _data_table_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Account", "Jan", "Feb"])
    ws.append(["Sales", 100, "=B2*2"])
    ws.append(["Payroll", 40, 44])
    return wb


def _new_project_conversation(client, name: str) -> tuple[str, str]:
    pid = client.post("/api/projects", json={"name": name}).json()["id"]
    cid = client.post(f"/api/projects/{pid}/conversations", json={"title": "Sheet drop"}).json()["id"]
    return pid, cid


def _upload(client, cid: str, wb: Workbook, filename: str = "book.xlsx") -> str:
    r = client.post(f"/api/conversations/{cid}/attachments",
                    files={"file": (filename, _xlsx_bytes(wb),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201
    return r.json()["id"]


def _send(client, cid: str, text: str, attachments: list[str] | None = None) -> str:
    body: dict = {"content": text}
    if attachments:
        body["attachments"] = attachments
    return client.post(f"/api/conversations/{cid}/messages", json=body).text


def _events(sse_text: str) -> list[tuple[str, dict]]:
    return [(m.group(1), json.loads(m.group(2)))
            for m in re.finditer(r"event: (\w+)\ndata: (.*?)\n\n", sse_text, re.S)]


def _blocks(sse_text: str) -> list[dict]:
    return [data for name, data in _events(sse_text) if name == "block"]


def _block_types(sse_text: str) -> list[str]:
    return [b["type"] for b in _blocks(sse_text)]


def _first_block(sse_text: str, block_type: str) -> dict:
    return next(b for b in _blocks(sse_text) if b["type"] == block_type)


def _markdown_text(sse_text: str) -> str:
    return "\n".join(b["data"].get("text", "") for b in _blocks(sse_text) if b["type"] == "markdown")


def _action_values(sse_text: str) -> list[str]:
    return [a["value"] for b in _blocks(sse_text) if b["type"] == "confirmation"
            for a in b["data"]["actions"]]


def _coa_turn(client, cid: str, attachment_id: str) -> str:
    return _send(client, cid, "here is our chart of accounts", attachments=[attachment_id])


# --- routing & registration --------------------------------------------------


def test_spreadsheet_skill_registered_and_in_catalog(client):
    assert "spreadsheet" in SKILLS
    assert len(SKILLS) == 14
    assert SKILLS["spreadsheet"].spec.approval_required is True

    names = {s["name"] for s in client.get("/api/skills").json()["skills"]}
    assert "/spreadsheet" in names
    catalog = {c["name"]: c for c in client.get("/api/meta/skills").json()["skills"]}
    entry = catalog["spreadsheet"]
    assert entry["title"] == "Spreadsheet Import"
    assert any("layout" in e for e in entry["examples"])
    assert any("macros" in e for e in entry["examples"])


def test_intent_routing_for_spreadsheet():
    assert detect_intent("/spreadsheet").skill == "spreadsheet"
    assert detect_intent("/excel what is in it").skill == "spreadsheet"
    assert detect_intent("/sheet").skill == "spreadsheet"
    assert detect_intent("what did you find in the spreadsheet I uploaded?").skill == "spreadsheet"
    assert detect_intent("look at the file I attached").skill == "spreadsheet"
    assert detect_intent("import this chart of accounts").skill == "spreadsheet"
    # unrelated phrasings keep their existing routes
    assert detect_intent("create an Actuals form").skill == "forms"
    assert detect_intent("build the context").skill == "context"


# --- analyze phase -----------------------------------------------------------


def test_attachment_message_streams_preview_and_links_attachment(client, session):
    _pid, cid = _new_project_conversation(client, "Sheet Analyze")
    att_id = _upload(client, cid, _coa_workbook(), "coa.xlsx")
    sse = _coa_turn(client, cid, att_id)

    types = _block_types(sse)
    assert "spreadsheetPreview" in types
    assert "markdown" in types
    assert "confirmation" in types

    preview = _first_block(sse, "spreadsheetPreview")["data"]
    assert preview["filename"] == "coa.xlsx"
    assert preview["sheetName"] == "Accounts"
    assert preview["kind"] == "chartOfAccounts"
    assert preview["memberCount"] == 5  # every row parses; existence is checked at validate time
    assert preview["dimensionGuess"] == "Account"
    assert [c["role"] for c in preview["columns"]] == ["member", "parent", "alias"]
    assert preview["sampleRows"][0][0] == "Total Payroll"
    assert isinstance(preview["issues"], list)

    values = _action_values(sse)
    assert "merge into context" in values
    assert "generate metadata csv" in values
    assert "validate against tenant" in values
    assert "explain formulas" not in values  # no formulas in this workbook
    assert "cancel" in values

    # the attachment row is linked to the persisted user message
    msgs = client.get(f"/api/conversations/{cid}/messages").json()
    user_msg = next(m for m in msgs if m["role"] == "user")
    session.expire_all()
    att = session.get(Attachment, att_id)
    assert att.message_id == user_msg["id"]
    assert att.conversation_id == cid


def test_unknown_sheet_is_honest(client):
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"
    ws.append(["Notes"])
    ws.append(["free text about the project"])
    ws.append(["more free text"])
    _pid, cid = _new_project_conversation(client, "Sheet Unknown")
    att_id = _upload(client, cid, wb, "notes.xlsx")
    sse = _send(client, cid, "what is this?", attachments=[att_id])
    assert "spreadsheetPreview" in _block_types(sse)
    md = _markdown_text(sse)
    assert "could not classify" in md
    assert "no import actions are available" in md


# --- CoA: merge into context -------------------------------------------------


def test_merge_into_context_creates_and_activates_version(client, session):
    pid, cid = _new_project_conversation(client, "Sheet Merge")
    att_id = _upload(client, cid, _coa_workbook(), "coa.xlsx")
    _coa_turn(client, cid, att_id)

    sse = _send(client, cid, "merge into context")
    types = _block_types(sse)
    assert "diff" in types
    md = _markdown_text(sse)
    assert "Imported_Account_1" in md
    assert "now active" in md

    session.expire_all()
    cv = context_store.get_active_context(session, pid)
    assert cv is not None
    assert cv.label == "Imported_Account_1"
    assert cv.mode == "imported"
    assert cv.active is True
    names = {r.name for r in context_store.get_records(session, cv.id, kind="member")}
    assert {"Total Payroll", "Salaries", "wages", "Overtime", "Rocket Fuel"} <= names

    diff = _first_block(sse, "diff")["data"]
    assert diff["before"] == "members: 0"
    assert diff["after"] == "members: 5"


# --- CoA: metadata CSV -------------------------------------------------------


def test_generate_metadata_csv_saves_artifact(client):
    pid, cid = _new_project_conversation(client, "Sheet CSV")
    att_id = _upload(client, cid, _coa_workbook(), "coa.xlsx")
    _coa_turn(client, cid, att_id)

    sse = _send(client, cid, "generate metadata csv")
    dl = _first_block(sse, "downloadableFile")["data"]
    assert dl["filename"] == "Account_metadata.csv"
    assert dl["checksum"]
    assert dl["sizeBytes"] > 0
    md = _markdown_text(sse)
    assert "importMetadata" in md
    assert "never executes" in md

    arts = client.get(f"/api/projects/{pid}/artifacts").json()
    csv_art = next(a for a in arts if a["kind"] == "metadataCsv")
    assert csv_art["name"] == "Account_metadata.csv"
    assert csv_art["id"] == dl["artifactId"]


# --- CoA: validate against tenant --------------------------------------------


def test_validate_against_tenant_reports_matches_and_missing(client):
    pid, cid = _new_project_conversation(client, "Sheet Validate")
    att_id = _upload(client, cid, _coa_workbook(), "coa.xlsx")
    _coa_turn(client, cid, att_id)

    sse = _send(client, cid, "validate against tenant")
    md = _markdown_text(sse)
    # 4 of 5 exist in MCWPCF Account: 3 exact + 1 case-insensitive ('wages' -> 'Wages')
    assert "**4 of 5** spreadsheet members exist" in md
    assert "exact: 3" in md
    assert "case-insensitive: 1" in md
    assert "**1 missing**" in md
    assert "| Rocket Fuel | missing |" in md
    assert "| wages | caseInsensitive | Wages |" in md
    # 'Overtime' has parent 'Total Revenue' in the sheet but 'Total Payroll' in the tenant
    assert "| Overtime | exact | Overtime | Total Revenue | Total Payroll | no |" in md

    dl = _first_block(sse, "downloadableFile")["data"]
    assert dl["filename"] == "Account_reconciliation.csv"
    arts = client.get(f"/api/projects/{pid}/artifacts").json()
    rec = next(a for a in arts if a["kind"] == "reconciliationCsv")
    assert rec["id"] == dl["artifactId"]


# --- layout -> form handoff --------------------------------------------------


def test_layout_to_form_handoff_and_follow_up_edit(client):
    _pid, cid = _new_project_conversation(client, "Sheet Layout")
    att_id = _upload(client, cid, _layout_workbook(), "layout.xlsx")
    sse = _send(client, cid, "I sketched a form", attachments=[att_id])
    preview = _first_block(sse, "spreadsheetPreview")["data"]
    assert preview["kind"] == "layout"
    values = _action_values(sse)
    assert "create a form from this layout" in values
    assert "create a report from this layout" in values

    sse = _send(client, cid, "create a form from this layout")
    types = _block_types(sse)
    assert "formPreview" in types
    assert "formSpecification" in types
    assert "validationReport" in types
    assert "confirmation" in types

    spec = _first_block(sse, "formSpecification")["data"]["spec"]
    rows = spec["rows"][0]
    assert rows["dimension"] == "Account"
    assert rows["selection"]["type"] == "memberList"
    assert rows["selection"]["members"] == ["Total Payroll", "Salaries", "Wages"]
    cols = spec["columns"][0]
    assert cols["dimension"] == "Period"
    assert cols["selection"]["type"] == "range"
    assert cols["selection"]["start"] == "Jan" and cols["selection"]["end"] == "Mar"
    pov_dims = {p["dimension"]: p for p in spec["pov"]}
    assert pov_dims["Scenario"]["selection"]["member"] == "Actual"
    # the unresolvable label is reported, never guessed
    md = _markdown_text(sse)
    assert "Rocket Fuel" in md
    assert "no exact name or alias match" in md

    # follow-up edit lands in the forms workflow (seamless handoff)
    sse = _send(client, cid, "hide Mar")
    assert "diff" in _block_types(sse)
    assert "formPreview" in _block_types(sse)


def test_layout_to_report_handoff(client):
    _pid, cid = _new_project_conversation(client, "Sheet Report")
    att_id = _upload(client, cid, _layout_workbook(), "layout.xlsx")
    _send(client, cid, "layout attached", attachments=[att_id])

    sse = _send(client, cid, "create a report from this layout")
    types = _block_types(sse)
    assert "reportPreview" in types
    assert "reportSpecification" in types
    spec = _first_block(sse, "reportSpecification")["data"]["spec"]
    grid = spec["grids"][0]
    assert grid["rows"][0]["selection"]["members"] == ["Total Payroll", "Salaries", "Wages"]
    assert grid["columns"][0]["selection"]["start"] == "Jan"


# --- data table --------------------------------------------------------------


def test_data_table_offers_generation_only_load_plan(client):
    _pid, cid = _new_project_conversation(client, "Sheet Data")
    att_id = _upload(client, cid, _data_table_workbook(), "actuals.xlsx")
    sse = _send(client, cid, "actuals data attached", attachments=[att_id])
    preview = _first_block(sse, "spreadsheetPreview")["data"]
    assert preview["kind"] == "dataTable"
    md = _markdown_text(sse)
    assert "generation only" in md
    values = _action_values(sse)
    assert "generate load-file plan" in values

    sse = _send(client, cid, "generate load-file plan")
    md = _markdown_text(sse)
    assert "importData" in md
    assert "uploadFile" in md
    assert "never executes" in md
    assert "generation only" in md


# --- formulas ----------------------------------------------------------------


def test_explain_formulas_emits_code_and_provider_prose(client):
    _pid, cid = _new_project_conversation(client, "Sheet Formulas")
    att_id = _upload(client, cid, _data_table_workbook(), "actuals.xlsx")
    _send(client, cid, "data attached", attachments=[att_id])

    sse = _send(client, cid, "explain formulas")
    types = _block_types(sse)
    assert "code" in types
    code = _first_block(sse, "code")["data"]["code"]
    assert "C2: =B2*2" in code
    # the mock provider streams deterministic prose after the code blocks
    tokens = "".join(d.get("text", "") for name, d in _events(sse) if name == "token")
    assert tokens.strip() != ""


# --- workflow hygiene --------------------------------------------------------


def test_cancel_ends_spreadsheet_workflow(client):
    _pid, cid = _new_project_conversation(client, "Sheet Cancel")
    att_id = _upload(client, cid, _coa_workbook(), "coa.xlsx")
    _coa_turn(client, cid, att_id)
    sse = _send(client, cid, "cancel")
    assert "Cancelled" in _markdown_text(sse)
    # after cancelling, a plain message no longer hits the spreadsheet workflow
    sse = _send(client, cid, "what cubes and dimensions exist?")
    assert "merge into context" not in sse
