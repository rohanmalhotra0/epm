"""Spreadsheet ingestion tests: deterministic parser, CoA hierarchy building,
metadata CSV rendering, context merge and the attachments API.

Real .xlsx files are built in-test with openpyxl. VBA extraction is unit-tested
against a stubbed oletools ``VBA_Parser`` — hand-crafting a real, valid
vbaProject.bin OLE binary is out of scope here, so the wrapper behaviour
(extraction caps, redaction, no execution, error containment) is what is
exercised.
"""

from __future__ import annotations

import hashlib
import io
from pathlib import Path

from openpyxl import Workbook

from app.services import context_store
from app.services import projects as projects_svc
from app.spreadsheet import analyze_file
from app.spreadsheet import parser as sheet_parser
from app.spreadsheet.context_merge import merge_hierarchy_into_context
from app.spreadsheet.metadata_csv import render_metadata_csv, save_metadata_artifact
from app.spreadsheet.models import HierarchyParse, ParsedMember

# --- workbook builders -------------------------------------------------------


def _coa_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    for row in (
        ["Account", "Parent", "Alias"],
        ["Income", "", "All Income"],
        ["Sales", "Income", "Product Sales"],
        ["Services", "Income", ""],
        ["Expenses", "", ""],
        ["Payroll", "Expenses", "Salaries"],
    ):
        ws.append(row)
    return wb


def _save(tmp_path: Path, wb: Workbook, name: str = "book.xlsx") -> Path:
    path = tmp_path / name
    wb.save(path)
    return path


def _xlsx_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- parser: chart of accounts ----------------------------------------------


def test_coa_parent_child_parsing(tmp_path):
    analysis = analyze_file(_save(tmp_path, _coa_workbook()))
    assert analysis.kind_guess == "chartOfAccounts"
    sheet = analysis.sheets[0]
    assert sheet.name == "Accounts"
    assert sheet.kind == "chartOfAccounts"
    h = sheet.hierarchy
    assert h is not None
    assert h.dimension_guess == "Account"
    assert [m.name for m in h.members] == ["Income", "Sales", "Services", "Expenses", "Payroll"]
    by_name = {m.name: m for m in h.members}
    assert by_name["Sales"].parent == "Income"
    assert by_name["Sales"].alias == "Product Sales"  # alias column detected
    assert by_name["Income"].parent is None
    assert h.root_count == 2
    assert h.issues == []


def test_coa_level_column_parsing(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Entity"
    for row in (
        ["Level 1", "Level 2", "Level 3", "Alias"],
        ["Income", None, None, "All Income"],
        [None, "Sales", None, None],
        [None, None, "Hardware", "HW"],
        [None, None, "Software", None],
        [None, "Services", None, None],
        ["Expenses", None, None, None],
    ):
        ws.append(row)
    sheet = analyze_file(_save(tmp_path, wb)).sheets[0]
    assert sheet.kind == "chartOfAccounts"
    h = sheet.hierarchy
    by_name = {m.name: m for m in h.members}
    assert by_name["Sales"].parent == "Income"
    assert by_name["Hardware"].parent == "Sales"
    assert by_name["Hardware"].alias == "HW"
    assert by_name["Software"].parent == "Sales"
    assert by_name["Services"].parent == "Income"
    assert by_name["Expenses"].parent is None
    assert h.root_count == 2
    assert h.dimension_guess == "Entity"  # level headers are generic; sheet name wins


def test_coa_invalid_identifier_recorded_as_issue(tmp_path):
    wb = Workbook()
    ws = wb.active
    for row in (
        ["Member", "Parent"],
        ["Income", ""],
        ["Bad;Name", "Income"],  # shell metacharacter — connector identifier rules reject it
    ):
        ws.append(row)
    h = analyze_file(_save(tmp_path, wb)).sheets[0].hierarchy
    assert [m.name for m in h.members] == ["Income"]
    assert any("disallowed character" in i for i in h.issues)


def test_coa_duplicate_orphan_and_cycle_issues(tmp_path):
    wb = Workbook()
    ws = wb.active
    for row in (
        ["Member", "Parent"],
        ["A", "B"],
        ["B", "A"],  # cycle A <-> B
        ["A", ""],  # duplicate
        ["C", "Missing"],  # orphan parent
    ):
        ws.append(row)
    h = analyze_file(_save(tmp_path, wb)).sheets[0].hierarchy
    assert any("duplicate member 'A'" in i for i in h.issues)
    assert any("unknown parent 'Missing'" in i for i in h.issues)
    assert any("cycle detected" in i for i in h.issues)


# --- parser: layout / data table / mess -------------------------------------


def test_layout_sheet_detection(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "FormLayout"
    ws.append(["Scenario: Actual"])
    ws.append([None, "Jan", "Feb", "Mar", "Q1"])
    ws.append(["Sales", 100, 110, 120, 330])
    ws.append(["Marketing", 50, 55, 60, 165])
    sheet = analyze_file(_save(tmp_path, wb)).sheets[0]
    assert sheet.kind == "layout"
    assert sheet.layout is not None
    assert sheet.layout.row_labels == ["Sales", "Marketing"]
    assert sheet.layout.column_labels == ["Jan", "Feb", "Mar", "Q1"]
    assert sheet.layout.pov_hints == ["Scenario: Actual"]


def test_data_table_detection_and_formulas(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Account", "Jan", "Feb"])
    ws.append(["Sales", 100, "=B2*2"])
    ws.append(["Payroll", 40, 44])
    sheet = analyze_file(_save(tmp_path, wb)).sheets[0]
    assert sheet.kind == "dataTable"
    table = sheet.data_table
    assert table.period_columns == ["Jan", "Feb"]
    assert table.label_column == "Account"
    assert table.row_count == 2
    # formulas are extracted as inert strings, never evaluated
    assert [(f.cell, f.formula) for f in sheet.formulas] == [("C2", "=B2*2")]


def test_messy_sheet_issues_recorded_no_crash(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Messy"
    ws.append(["Corporate Chart of Accounts"])
    ws.merge_cells("A1:C1")
    ws.append(["Account", "Parent", "Alias"])
    ws.append(["Income", "", ""])
    ws.append(["Sales", "Income", ""])
    ws.append(["Total Expenses", "", ""])
    sheet = analyze_file(_save(tmp_path, wb)).sheets[0]
    assert sheet.kind == "chartOfAccounts"  # still classified despite the mess
    assert any("merged cell" in i for i in sheet.issues)
    assert any("title/banner" in i for i in sheet.issues)
    assert any("subtotal" in i for i in sheet.issues)
    assert {m.name for m in sheet.hierarchy.members} == {"Income", "Sales", "Total Expenses"}


def test_csv_parsing(tmp_path):
    path = tmp_path / "coa.csv"
    path.write_text("Account,Parent,Alias\nIncome,,All Income\nSales,Income,Product Sales\n")
    analysis = analyze_file(path)
    assert analysis.kind_guess == "chartOfAccounts"
    assert analysis.sheets[0].name == "coa"
    assert analysis.sheets[0].hierarchy.members[1].alias == "Product Sales"


def test_unclassifiable_sheet_is_honest(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Notes"])
    ws.append(["free text about the project"])
    ws.append(["more free text"])
    sheet = analyze_file(_save(tmp_path, wb)).sheets[0]
    assert sheet.kind == "unknown"
    assert any("could not classify" in i for i in sheet.issues)


# --- VBA extraction (stubbed — see module docstring) -------------------------


class _StubVbaParser:
    def __init__(self, _path: str) -> None:
        pass

    def detect_vba_macros(self) -> bool:
        return True

    def extract_macros(self):
        yield ("book.xlsm", "VBA/Module1", "Module1",
               'Sub Main()\n  password = "hunter2secret"\nEnd Sub')
        yield ("book.xlsm", "VBA/Module2", "Module2", "x" * (sheet_parser.MAX_VBA_CHARS + 500))

    def close(self) -> None:
        pass


def test_vba_extraction_stub_redacts_and_caps(monkeypatch, tmp_path):
    from oletools import olevba

    monkeypatch.setattr(olevba, "VBA_Parser", _StubVbaParser)
    modules, issues = sheet_parser.extract_vba_modules(tmp_path / "book.xlsm")
    assert [m.name for m in modules] == ["Module1", "Module2"]
    assert "hunter2secret" not in modules[0].code  # redact_text applied
    assert "«redacted»" in modules[0].code
    assert modules[0].line_count == 3
    assert any("truncated" in i for i in issues)
    assert len(modules[1].code) <= sheet_parser.MAX_VBA_CHARS


def test_vba_extraction_failure_contained(monkeypatch, tmp_path):
    from oletools import olevba

    class Exploding:
        def __init__(self, _path: str) -> None:
            raise ValueError("corrupt OLE stream")

    monkeypatch.setattr(olevba, "VBA_Parser", Exploding)
    modules, issues = sheet_parser.extract_vba_modules(tmp_path / "book.xlsm")
    assert modules == []
    assert any("VBA extraction failed" in i for i in issues)


# --- metadata CSV ------------------------------------------------------------


def test_metadata_csv_content_and_determinism(tmp_path):
    path = _save(tmp_path, _coa_workbook())
    h1 = analyze_file(path).sheets[0].hierarchy
    h2 = analyze_file(path).sheets[0].hierarchy
    csv1 = render_metadata_csv(h1, "Account")
    csv2 = render_metadata_csv(h2, "Account")
    assert csv1.encode() == csv2.encode()  # same input -> identical bytes
    assert csv1 == (
        "Account,Parent,Alias: Default,Data Storage\n"
        "Income,Account,All Income,Store\n"
        "Sales,Income,Product Sales,Store\n"
        "Services,Income,,Store\n"
        "Expenses,Account,,Store\n"
        "Payroll,Expenses,Salaries,Store\n"
    )


def test_save_metadata_artifact(session):
    project = projects_svc.create_project(session, "Metadata CSV Project")
    h = HierarchyParse(
        dimension_guess="Account", root_count=1,
        members=[ParsedMember(name="Income"), ParsedMember(name="Sales", parent="Income")],
    )
    artifact = save_metadata_artifact(session, project.id, h, "Account")
    assert artifact.kind == "metadataCsv"
    assert artifact.name == "Account_metadata.csv"
    assert artifact.content.startswith("Account,Parent,Alias: Default,Data Storage\n")
    assert artifact.checksum == hashlib.sha256(artifact.content.encode("utf-8")).hexdigest()
    assert artifact.metadata_["memberCount"] == 2


# --- context merge -----------------------------------------------------------


def test_context_merge_creates_and_activates_new_versions(session):
    project = projects_svc.create_project(session, "Merge Project")
    h1 = HierarchyParse(
        dimension_guess="Account", root_count=1,
        members=[
            ParsedMember(name="Income"),
            ParsedMember(name="Sales", parent="Income", alias="Product Sales"),
        ],
    )
    cv1 = merge_hierarchy_into_context(session, project.id, h1, "Account")
    session.flush()
    assert cv1.label == "Imported_Account_1"
    assert cv1.mode == "imported"
    assert cv1.active is True
    records = context_store.get_records(session, cv1.id, kind="member")
    assert {r.name for r in records} == {"Income", "Sales"}
    sales = next(r for r in records if r.name == "Sales")
    assert sales.parent == "Income"
    assert sales.alias == "Product Sales"
    assert sales.dimension == "Account"
    assert sales.search_text == "sales product sales"
    assert sales.data["name"] == "Sales"

    # second import: a NEW version copies the first and activates; v1 untouched
    h2 = HierarchyParse(dimension_guess="Account", root_count=1,
                        members=[ParsedMember(name="Other")])
    cv2 = merge_hierarchy_into_context(session, project.id, h2, "Account")
    session.flush()
    session.expire_all()
    assert cv2.id != cv1.id
    assert cv2.label == "Imported_Account_2"
    assert cv2.active is True
    assert cv1.active is False
    assert len(context_store.get_records(session, cv1.id)) == 2  # prior version untouched
    assert {r.name for r in context_store.get_records(session, cv2.id)} == {"Income", "Sales", "Other"}
    assert session.get(type(cv2), cv2.id).counts["members"] == 3
    proj = projects_svc.get_project(session, project.id)
    assert proj.active_context_version_id == cv2.id


# --- attachments API ---------------------------------------------------------


def _conversation_id(client) -> str:
    pid = client.get("/api/projects").json()[0]["id"]
    return client.post(f"/api/projects/{pid}/conversations", json={"title": "Sheet drop"}).json()["id"]


def test_upload_attachment_201_shape(client):
    cid = _conversation_id(client)
    data = _xlsx_bytes(_coa_workbook())
    r = client.post(
        f"/api/conversations/{cid}/attachments",
        files={"file": ("My CoA.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 201
    body = r.json()
    assert set(body) == {"id", "conversationId", "projectId", "filename", "mediaType",
                        "sizeBytes", "checksum", "sheetNames", "kindGuess"}
    assert body["conversationId"] == cid
    assert body["filename"] == "My CoA.xlsx"
    assert body["mediaType"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert body["sizeBytes"] == len(data)
    assert body["checksum"] == hashlib.sha256(data).hexdigest()
    assert body["sheetNames"] == ["Accounts"]
    assert body["kindGuess"] == "chartOfAccounts"

    # GET round-trips the same shape
    got = client.get(f"/api/attachments/{body['id']}")
    assert got.status_code == 200
    assert got.json() == body


def test_upload_attachment_stores_redacted_text_extract(client, session):
    from app.db.models import Attachment

    cid = _conversation_id(client)
    r = client.post(f"/api/conversations/{cid}/attachments",
                    files={"file": ("coa.xlsx", _xlsx_bytes(_coa_workbook()), "application/octet-stream")})
    att = session.get(Attachment, r.json()["id"])
    assert att is not None
    assert "Accounts" in att.text_extract and "Parent" in att.text_extract
    assert Path(att.path).exists()
    assert Path(att.path).parent.name == att.id


def test_upload_rejects_bad_extension(client):
    cid = _conversation_id(client)
    r = client.post(f"/api/conversations/{cid}/attachments",
                    files={"file": ("notes.txt", b"hello", "text/plain")})
    assert r.status_code == 400
    assert "unsupported file type" in r.json()["detail"]


def test_upload_rejects_oversize_file(client):
    cid = _conversation_id(client)
    big = b"a" * (10 * 1024 * 1024 + 1)
    r = client.post(f"/api/conversations/{cid}/attachments",
                    files={"file": ("big.csv", big, "text/csv")})
    assert r.status_code == 400
    assert "10 MB" in r.json()["detail"]


def test_upload_unknown_conversation_404(client):
    r = client.post("/api/conversations/nope/attachments",
                    files={"file": ("coa.csv", b"Account,Parent\nIncome,\n", "text/csv")})
    assert r.status_code == 404


def test_attachment_analysis_endpoint(client):
    cid = _conversation_id(client)
    att = client.post(f"/api/conversations/{cid}/attachments",
                      files={"file": ("coa.xlsx", _xlsx_bytes(_coa_workbook()),
                                      "application/octet-stream")}).json()
    r = client.get(f"/api/attachments/{att['id']}/analysis")
    assert r.status_code == 200
    analysis = r.json()
    assert analysis["filename"] == "coa.xlsx"
    assert analysis["kindGuess"] == "chartOfAccounts"
    assert analysis["vbaModules"] == []
    sheet = analysis["sheets"][0]
    assert sheet["kind"] == "chartOfAccounts"
    assert [c["role"] for c in sheet["columns"]] == ["member", "parent", "alias"]
    hierarchy = sheet["hierarchy"]
    assert hierarchy["dimensionGuess"] == "Account"
    assert hierarchy["rootCount"] == 2
    assert hierarchy["members"][1] == {
        "name": "Sales", "parent": "Income", "alias": "Product Sales", "storage": None,
    }
    assert sheet["sampleRows"][0][0] == "Income"

    # unknown attachment
    assert client.get("/api/attachments/nope/analysis").status_code == 404


def test_upload_csv_attachment(client):
    cid = _conversation_id(client)
    r = client.post(
        f"/api/conversations/{cid}/attachments",
        files={"file": ("entity_load.csv", b"Entity,Parent\nTotal Entity,\nUS,Total Entity\n", "text/csv")},
    )
    assert r.status_code == 201
    body = r.json()
    assert body["mediaType"] == "text/csv"
    assert body["kindGuess"] == "chartOfAccounts"
    assert body["sheetNames"] == ["entity_load"]
