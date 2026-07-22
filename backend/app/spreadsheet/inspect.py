"""Deterministic workbook *inspection* — the "see everything" surface.

Where :mod:`parser` classifies a workbook for the chat ingestion flow, this
module answers a different question: *what's inside and what makes it move?* It
surfaces every VBA macro's source, every named range, table, pivot, chart and
external data connection, and — importantly — the **auto-run hooks**
(``Workbook_Open``, ``Auto_Open``, ``Worksheet_Change`` …) that let a workbook
act on its own.

Parse-only, like everything else in this package: nothing is compiled,
evaluated or executed. VBA is extracted as inert, redacted text; connection
strings and command text are redacted; formulas are never evaluated.
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from ..security.redaction import redact_text
from .models import (
    ChartInfo,
    DataConnection,
    NamedRange,
    PivotTableInfo,
    SheetSummary,
    TableInfo,
    VbaModule,
    VbaProcedure,
    WorkbookInspection,
    WorkbookTrigger,
)
from .parser import analyze_file, extract_vba_modules

_SS_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS = {"m": _SS_NS}

# OOXML connection type codes (SpreadsheetML @type on <connection>).
_CONN_TYPES = {
    "1": "OLE DB", "2": "DAO", "3": "File", "4": "Web", "5": "ODBC",
    "6": "Text", "7": "OLAP", "8": "Worksheet",
}

_SUPPORTED = {".xlsx", ".xlsm", ".xlsb", ".csv"}


# --- VBA procedure / trigger scanning (pure — unit-tested without oletools) --

_PROC_RE = re.compile(
    r"^[ \t]*(?:Public|Private|Friend|Global)?[ \t]*(?:Static[ \t]+)?"
    r"(Sub|Function|Property[ \t]+(?:Get|Let|Set))[ \t]+([A-Za-z_][A-Za-z0-9_]*)"
    r"[ \t]*(\([^\r\n)]*\))?",
    re.IGNORECASE | re.MULTILINE,
)

# Classic Auto_* entry points Excel runs without any object event.
_AUTO_NAMES = {"auto_open", "auto_close", "auto_exec", "auto_activate", "auto_deactivate"}
# Host-object event handlers (Workbook_Open, Worksheet_Change, …).
_EVENT_PREFIXES = ("workbook_", "worksheet_", "chart_", "app_", "application_")


def _trigger_scope(name: str) -> str | None:
    low = name.lower()
    if low in _AUTO_NAMES:
        return "auto"
    if low.startswith(_EVENT_PREFIXES):
        return "event"
    return None


def scan_vba_procedures(
    modules: list[VbaModule],
) -> tuple[list[VbaProcedure], list[WorkbookTrigger]]:
    """Extract Sub/Function/Property declarations and the auto-run hooks.

    Operates purely on the (already inert, already redacted) module text, so it
    needs no oletools and is safe to unit-test with plain strings.
    """
    procs: list[VbaProcedure] = []
    triggers: list[WorkbookTrigger] = []
    for mod in modules:
        for m in _PROC_RE.finditer(mod.code or ""):
            kind = re.sub(r"[ \t]+", " ", m.group(1)).strip().title().replace("Property ", "Property ")
            name = m.group(2)
            sig = (m.group(0) or "").strip()
            scope = _trigger_scope(name)
            procs.append(VbaProcedure(
                module=mod.name, name=name, kind=kind,
                signature=redact_text(sig)[:200], auto_run=scope is not None,
            ))
            if scope is not None:
                triggers.append(WorkbookTrigger(name=name, module=mod.name, scope=scope))
    procs.sort(key=lambda p: (p.module, p.name.lower()))
    triggers.sort(key=lambda t: t.name.lower())
    return procs, triggers


# --- openpyxl structure (sheets, names, tables, charts) ---------------------

def _chart_title(chart: object) -> str:
    try:
        rich = chart.title.tx.rich  # type: ignore[attr-defined]
        parts = [r.t for p in rich.p for r in p.r if getattr(r, "t", None)]
        return "".join(parts).strip()
    except Exception:
        return ""


def _openpyxl_structure(path: Path) -> tuple[dict, list[NamedRange], list[TableInfo], list[ChartInfo], list[str]]:
    """Per-sheet metadata + named ranges + tables + charts, via openpyxl.

    Returns ({sheet_name: {...}}, named_ranges, tables, charts, issues).
    Raises are caller-handled (e.g. .xlsb is not readable by openpyxl).
    """
    from openpyxl import load_workbook

    issues: list[str] = []
    sheet_meta: dict[str, dict] = {}
    named: list[NamedRange] = []
    tables: list[TableInfo] = []
    charts: list[ChartInfo] = []

    wb = load_workbook(path, data_only=False, read_only=False, keep_vba=False, keep_links=False)
    try:
        for idx, ws in enumerate(wb.worksheets):
            t_count = 0
            # NB: TableList.items()/.values() yield the ref STRING, not the Table
            # object — only __getitem__ resolves to the Table. Index by name.
            for tname in list(getattr(ws, "tables", {}) or {}):
                tbl = ws.tables[tname]
                cols: list[str] = []
                try:
                    cols = [c.name for c in (tbl.tableColumns or []) if getattr(c, "name", None)]
                except Exception:
                    pass
                tables.append(TableInfo(name=tname, sheet=ws.title,
                                        ref=getattr(tbl, "ref", "") or "", columns=cols))
                t_count += 1
            c_list = list(getattr(ws, "_charts", []) or [])
            for ch in c_list:
                charts.append(ChartInfo(sheet=ws.title,
                                        chart_type=type(ch).__name__, title=_chart_title(ch)))
            # Sheet-scoped defined names (openpyxl 3.1 puts these on the sheet).
            try:
                for nm, dn in (getattr(ws, "defined_names", {}) or {}).items():
                    if nm.startswith("_xlnm"):
                        continue
                    named.append(NamedRange(name=nm, refers_to=(dn.value or ""),
                                            scope=ws.title, hidden=bool(getattr(dn, "hidden", False))))
            except Exception:
                pass
            sheet_meta[ws.title] = {
                "index": idx,
                "visibility": getattr(ws, "sheet_state", "visible") or "visible",
                "dimensions": ws.dimensions or "",
                "rows": int(ws.max_row or 0),
                "cols": int(ws.max_column or 0),
                "table_count": t_count,
                "chart_count": len(c_list),
            }
        # Workbook-scoped defined names.
        try:
            for nm, dn in (wb.defined_names or {}).items():
                if nm.startswith("_xlnm"):
                    continue
                named.append(NamedRange(name=nm, refers_to=(dn.value or ""),
                                        scope="workbook", hidden=bool(getattr(dn, "hidden", False))))
        except Exception:
            pass
    finally:
        wb.close()
    named.sort(key=lambda n: n.name.lower())
    tables.sort(key=lambda t: (t.sheet, t.name.lower()))
    return sheet_meta, named, tables, charts, issues


# --- OOXML zip parts (pivots, connections, chart fallback) ------------------

def _ooxml_parts(path: Path) -> tuple[list[PivotTableInfo], list[DataConnection], int, bool, list[str]]:
    """Parse pivot tables, external connections and a chart-count fallback
    straight from the .xlsx/.xlsm zip. Returns
    (pivots, connections, chart_part_count, has_vba_project, issues)."""
    pivots: list[PivotTableInfo] = []
    conns: list[DataConnection] = []
    chart_parts = 0
    has_vba = False
    issues: list[str] = []
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            has_vba = "xl/vbaProject.bin" in names
            chart_parts = sum(1 for n in names if re.match(r"xl/charts/chart\d+\.xml$", n))

            if "xl/connections.xml" in names:
                try:
                    root = ET.fromstring(z.read("xl/connections.xml"))
                    for c in root.findall("m:connection", _NS):
                        db = c.find("m:dbPr", _NS)
                        web = c.find("m:webPr", _NS)
                        txt = c.find("m:textPr", _NS)
                        source = (db.get("connection") if db is not None else
                                  web.get("url") if web is not None else
                                  txt.get("sourceFile") if txt is not None else "") or ""
                        command = (db.get("command") if db is not None else "") or ""
                        conns.append(DataConnection(
                            name=c.get("name") or "",
                            type=_CONN_TYPES.get(c.get("type") or "", c.get("type") or ""),
                            source=redact_text(source)[:400],
                            command=redact_text(command)[:400],
                        ))
                except ET.ParseError:
                    issues.append("connections.xml was present but unparseable")

            cache_sources: list[str] = []
            for n in names:
                if re.match(r"xl/pivotCache/pivotCacheDefinition\d+\.xml$", n):
                    src = _pivot_cache_source(z, n)
                    if src:
                        cache_sources.append(src)
            for n in sorted(names):
                if re.match(r"xl/pivotTables/pivotTable\d+\.xml$", n):
                    try:
                        root = ET.fromstring(z.read(n))
                        loc = root.find("m:location", _NS)
                        pivots.append(PivotTableInfo(
                            name=root.get("name") or "PivotTable",
                            location=(loc.get("ref") if loc is not None else "") or "",
                            source="",
                        ))
                    except ET.ParseError:
                        issues.append(f"{n} was unparseable")
            # Best-effort source attribution: exact 1:1 pivot↔cache maps cleanly;
            # otherwise surface the distinct cache sources as a note.
            if pivots and len(cache_sources) == len(pivots):
                for p, s in zip(pivots, cache_sources):
                    p.source = s
            elif cache_sources:
                uniq = sorted(set(cache_sources))
                issues.append("pivot data source(s): " + "; ".join(uniq))
    except zipfile.BadZipFile:
        issues.append("file is not a valid Office Open XML (zip) container")
    return pivots, conns, chart_parts, has_vba, issues


def _pivot_cache_source(z: zipfile.ZipFile, part: str) -> str:
    try:
        root = ET.fromstring(z.read(part))
        cs = root.find("m:cacheSource", _NS)
        if cs is None:
            return ""
        if cs.get("type") == "external":
            return "(external connection)"
        wsrc = cs.find("m:worksheetSource", _NS)
        if wsrc is not None:
            sheet = wsrc.get("sheet") or ""
            ref = wsrc.get("ref") or wsrc.get("name") or ""
            return f"{sheet}!{ref}" if sheet and ref else (ref or sheet)
    except ET.ParseError:
        return ""
    return ""


# --- entry point ------------------------------------------------------------

def inspect_file(path: Path, filename: str | None = None, size_bytes: int = 0) -> WorkbookInspection:
    """Inspect a workbook file end to end. Parse only — never executes."""
    filename = filename or path.name
    ext = (Path(filename).suffix.lower() or path.suffix.lower())
    fmt = ext.lstrip(".")
    issues: list[str] = []

    if ext == ".csv":
        analysis = analyze_file(path, filename)
        s = analysis.sheets[0] if analysis.sheets else None
        summary = SheetSummary(
            name=s.name if s else "Sheet1", index=0,
            rows=len(s.sample_rows) if s else 0, cols=len(s.columns) if s else 0,
            formula_count=len(s.formulas) if s else 0, kind=s.kind if s else None,  # type: ignore[arg-type]
        ) if s else None
        return WorkbookInspection(
            filename=filename, size_bytes=size_bytes, file_format="csv",
            summary="CSV file · no macros or workbook structure",
            sheet_count=1 if summary else 0, sheets=[summary] if summary else [],
            issues=analysis.sheets[0].issues if analysis.sheets else [],
        )

    # xlsx / xlsm / xlsb — reuse the parser for sheets/formulas/VBA, then enrich.
    vba_modules: list[VbaModule] = []
    formula_by_sheet: dict[str, int] = {}
    kind_by_sheet: dict[str, object] = {}
    try:
        analysis = analyze_file(path, filename)
        for sh in analysis.sheets:
            formula_by_sheet[sh.name] = len(sh.formulas)
            kind_by_sheet[sh.name] = sh.kind
        vba_modules = list(analysis.vba_modules)
    except Exception as exc:  # noqa: BLE001 — .xlsb / corrupt files must not 500
        issues.append(f"structural parse limited: {redact_text(str(exc))[:200]}")

    # .xlsb never reaches openpyxl VBA extraction via analyze_file (it only runs
    # for .xlsm there); pull macros directly so .xlsb still shows its code.
    if not vba_modules and ext in (".xlsm", ".xlsb"):
        vba_modules, vba_issues = extract_vba_modules(path)
        issues.extend(vba_issues)

    sheet_meta: dict[str, dict] = {}
    named_ranges: list[NamedRange] = []
    tables: list[TableInfo] = []
    charts: list[ChartInfo] = []
    try:
        sheet_meta, named_ranges, tables, charts, struct_issues = _openpyxl_structure(path)
        issues.extend(struct_issues)
    except Exception as exc:  # noqa: BLE001
        issues.append(f"structure extraction limited ({fmt}): {redact_text(str(exc))[:160]}")

    pivots, connections, chart_parts, has_vba_bin, ooxml_issues = _ooxml_parts(path)
    issues.extend(ooxml_issues)
    if not charts and chart_parts:
        charts = [ChartInfo(sheet="", chart_type="Chart", title="") for _ in range(chart_parts)]

    procedures, triggers = scan_vba_procedures(vba_modules)

    # Assemble per-sheet summaries in workbook order (fall back to parser order).
    sheets: list[SheetSummary] = []
    ordered = sorted(sheet_meta.items(), key=lambda kv: kv[1]["index"]) if sheet_meta else \
        [(n, {"index": i}) for i, n in enumerate(formula_by_sheet)]
    tables_by_sheet: dict[str, int] = {}
    charts_by_sheet: dict[str, int] = {}
    for t in tables:
        tables_by_sheet[t.sheet] = tables_by_sheet.get(t.sheet, 0) + 1
    for c in charts:
        charts_by_sheet[c.sheet] = charts_by_sheet.get(c.sheet, 0) + 1
    for name, meta in ordered:
        sheets.append(SheetSummary(
            name=name, index=meta.get("index", 0),
            visibility=meta.get("visibility", "visible"),
            dimensions=meta.get("dimensions", ""),
            rows=meta.get("rows", 0), cols=meta.get("cols", 0),
            formula_count=formula_by_sheet.get(name, 0),
            table_count=meta.get("table_count", tables_by_sheet.get(name, 0)),
            chart_count=meta.get("chart_count", charts_by_sheet.get(name, 0)),
            kind=kind_by_sheet.get(name),  # type: ignore[arg-type]
        ))

    macro_enabled = ext in (".xlsm", ".xlsb") or has_vba_bin
    has_macros = bool(vba_modules)
    inspection = WorkbookInspection(
        filename=filename, size_bytes=size_bytes, file_format=fmt,
        macro_enabled=macro_enabled, has_macros=has_macros,
        sheet_count=len(sheets), sheets=sheets,
        vba_modules=vba_modules, procedures=procedures, triggers=triggers,
        named_ranges=named_ranges, tables=tables, pivot_tables=pivots,
        charts=charts, connections=connections, issues=issues,
    )
    inspection.summary = _summarize(inspection)
    return inspection


def _summarize(w: WorkbookInspection) -> str:
    bits = [f"{w.sheet_count} sheet{'s' if w.sheet_count != 1 else ''}"]
    if w.has_macros:
        bits.append(f"{len(w.procedures)} macro{'s' if len(w.procedures) != 1 else ''} "
                    f"in {len(w.vba_modules)} module{'s' if len(w.vba_modules) != 1 else ''}")
    elif w.macro_enabled:
        bits.append("macro-enabled but no VBA code found")
    if w.triggers:
        bits.append("auto-runs " + ", ".join(sorted({t.name for t in w.triggers}))[:80])
    if w.named_ranges:
        bits.append(f"{len(w.named_ranges)} named range{'s' if len(w.named_ranges) != 1 else ''}")
    if w.tables:
        bits.append(f"{len(w.tables)} table{'s' if len(w.tables) != 1 else ''}")
    if w.pivot_tables:
        bits.append(f"{len(w.pivot_tables)} pivot{'s' if len(w.pivot_tables) != 1 else ''}")
    if w.connections:
        bits.append(f"{len(w.connections)} data connection{'s' if len(w.connections) != 1 else ''}")
    if w.charts:
        bits.append(f"{len(w.charts)} chart{'s' if len(w.charts) != 1 else ''}")
    lead = "Macro-enabled workbook" if w.macro_enabled else "Workbook"
    return f"{lead} · " + " · ".join(bits)
