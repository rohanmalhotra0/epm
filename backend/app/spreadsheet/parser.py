"""Deterministic, parse-only spreadsheet analysis.

Reads .xlsx/.xlsm via openpyxl (``data_only=False`` so formulas are visible as
inert strings) and .csv via the stdlib ``csv`` module. Nothing is ever
executed: no macros, no formula evaluation. Real-world mess (merged cells,
title/multi-row headers, subtotal rows, truncation) is detected and recorded
in ``issues`` instead of silently guessed away.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import re
from pathlib import Path

from ..security.redaction import redact_text
from . import coa
from .models import (
    ColumnInfo,
    ColumnRole,
    DataTableParse,
    FormulaInfo,
    LayoutParse,
    SheetAnalysis,
    SheetKind,
    VbaModule,
    WorkbookAnalysis,
)

MAX_ROWS = 5000
# Daily cash sheets run a column per calendar day, so a year is ~260 columns plus
# label/source columns. 64 clipped those to a single quarter.
MAX_COLS = 400
MAX_FORMULAS_PER_SHEET = 200
MAX_SAMPLE_ROWS = 5
MAX_SAMPLE_COLS = 20
MAX_ISSUES_PER_KIND = 20
MAX_VBA_CHARS = 200_000

_MONTHS = {
    "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "sept", "oct", "nov", "dec",
    "january", "february", "march", "april", "june", "july", "august", "september",
    "october", "november", "december",
}
_PERIOD_RE = re.compile(
    r"^(q[1-4]|fy[ -]?\d{2,4}|(19|20)\d{2}|ytd|qtd|mtd|yeartotal|year ?total|p\d{1,2}|period ?\d{1,2})$",
    re.IGNORECASE,
)
# A date typed as text: 12/31/25, 2026-01-01, 1.2.26. Anchored and bounded so it
# can't swallow account codes like "4000.10".
_DATE_TEXT_RE = re.compile(
    r"^(?:\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4}|\d{4}[/\-.]\d{1,2}[/\-.]\d{1,2})$"
)
_LEVEL_RE = re.compile(r"^(level|gen(eration)?)[-_ ]?(\d+)$", re.IGNORECASE)
_TOTAL_RE = re.compile(r"(?i)(^(sub[- ]?)?total\b|\btotal$)")
_NUMBER_RE = re.compile(r"^\(?-?[\d,]+(\.\d+)?\)?%?$")

_MEMBER_HEADERS = {
    "member", "member name", "child", "child member", "account", "account name",
    "entity", "name", "dimension member",
}
_PARENT_HEADERS = {"parent", "parent member", "parent name"}
_ALIAS_HEADERS = {"description", "desc", "member description", "alias name", "member alias"}
_STORAGE_HEADERS = {"data storage", "storage", "data storage (account)"}


# --- small helpers ----------------------------------------------------------


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    # Render dates as dates. str(datetime) yields "2026-01-01 00:00:00", which is
    # noise in a column header and defeats the date grammar below.
    if isinstance(value, dt.datetime):
        return value.date().isoformat() if value.time() == dt.time.min else value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def _is_empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _is_numericish(value: object) -> bool:
    """Numbers, numeric-looking strings and formulas count as data-ish."""
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        t = value.strip()
        if t.startswith("="):
            return True  # a formula computes a value; treated as data, never evaluated
        t = t.replace("$", "").replace("€", "").replace("£", "").replace(" ", "")
        return bool(t) and bool(_NUMBER_RE.match(t))
    return False


def _is_period_header(text: str) -> bool:
    t = text.strip().lower().rstrip(".")
    if not t:
        return False
    if t in _MONTHS or _PERIOD_RE.match(t):
        return True
    if _DATE_TEXT_RE.match(t):  # "12/31/25", "2026-01-01", "1.2.26"
        return True
    # "Jan-24", "Jan 2025", "Oct FY24"
    first = re.split(r"[\s\-/]+", t)[0]
    return first in _MONTHS


def _is_period_cell(value: object) -> bool:
    """Period-ness of a raw cell, including real dates.

    Excel stores ``12/31/25`` as a date, so openpyxl hands back a ``datetime``
    however the cell is formatted. The old string-only test discarded those
    silently, which made every daily-column sheet unclassifiable.
    """
    if isinstance(value, bool):
        return False
    if isinstance(value, (dt.datetime, dt.date)):
        return True
    return isinstance(value, str) and _is_period_header(value)


def _nonempty_count(row: list[object]) -> int:
    return sum(1 for v in row if not _is_empty(v))


def _row_all_text(row: list[object]) -> bool:
    vals = [v for v in row if not _is_empty(v)]
    return bool(vals) and all(isinstance(v, str) and not _is_numericish(v) for v in vals)


def _stringify_row(row: list[object]) -> list[str]:
    return [_cell_text(v) for v in row[:MAX_SAMPLE_COLS]]


# --- header / column analysis -----------------------------------------------


def _header_role(header: str, member_found: bool) -> ColumnRole | None:
    h = header.strip().lower()
    if not h:
        return None
    if h in _PARENT_HEADERS:
        return ColumnRole.parent
    if _LEVEL_RE.match(h):
        return ColumnRole.level
    if h.startswith("alias") or h in _ALIAS_HEADERS:
        return ColumnRole.alias
    if h in _MEMBER_HEADERS:
        if not member_found:
            return ColumnRole.member
        return ColumnRole.alias if h in ("name", "member name") else ColumnRole.label
    if h in _STORAGE_HEADERS:
        return ColumnRole.label  # storage columns are located by header in coa.py
    if _is_period_header(h):
        return ColumnRole.period
    return None


def _body_role(col_values: list[object]) -> ColumnRole:
    vals = [v for v in col_values if not _is_empty(v)]
    if not vals:
        return ColumnRole.unknown
    numeric = sum(1 for v in vals if _is_numericish(v))
    if numeric / len(vals) >= 0.6:
        return ColumnRole.data
    return ColumnRole.label


def _build_columns(headers: list[str], body: list[list[object]]) -> list[ColumnInfo]:
    columns: list[ColumnInfo] = []
    member_found = False
    for idx, header in enumerate(headers):
        role = _header_role(header, member_found)
        if role == ColumnRole.member:
            member_found = True
        if role is None:
            role = _body_role([row[idx] if idx < len(row) else None for row in body])
        columns.append(ColumnInfo(index=idx, header=header, role=role))
    return columns


def _find_period_row(rows: list[list[object]], limit: int = 10) -> int | None:
    """Index of the densest period-header row, or None.

    Densest rather than first: a daily sheet stacks "Actual"/weekday/date rows,
    and weekday names would otherwise win on a sheet where they happen to parse.
    """
    best_idx, best_n = None, 1
    for idx, row in enumerate(rows[:limit]):
        n = sum(1 for v in row if _is_period_cell(v))
        if n > best_n:
            best_idx, best_n = idx, n
    return best_idx


def _detect_subtotal_rows(body: list[list[object]], header_rows: int, issues: list[str]) -> None:
    found = 0
    for i, row in enumerate(body):
        first_text = next((v for v in row if isinstance(v, str) and v.strip()), None)
        if first_text and _TOTAL_RE.search(first_text.strip()):
            found += 1
            if found <= MAX_ISSUES_PER_KIND:
                issues.append(
                    f"possible subtotal row at sheet row {header_rows + i + 1}: '{first_text.strip()}'"
                )
    if found > MAX_ISSUES_PER_KIND:
        issues.append(f"{found - MAX_ISSUES_PER_KIND} further possible subtotal rows not listed")


# --- per-sheet analysis -----------------------------------------------------


def _analyze_rows(
    sheet_name: str,
    rows: list[list[object]],
    formulas: list[FormulaInfo],
    issues: list[str],
) -> SheetAnalysis:
    # trim fully-empty trailing rows
    while rows and _nonempty_count(rows[-1]) == 0:
        rows.pop()
    if not rows or all(_nonempty_count(r) == 0 for r in rows):
        issues.append("sheet is empty")
        return SheetAnalysis(name=sheet_name, kind=SheetKind.unknown, formulas=formulas, issues=issues)

    width = max(len(r) for r in rows)
    rows = [list(r) + [None] * (width - len(r)) for r in rows]

    first_idx = next(i for i, r in enumerate(rows) if _nonempty_count(r) > 0)
    header_idx = first_idx
    header_row = rows[header_idx]
    combined = False

    # Skip stacked banner rows. Real sheets stack more than one (a title, then a
    # subtitle), so this repeats rather than firing once; bounded so a sheet of
    # single-cell rows can't consume the whole body.
    banner_from = header_idx
    while header_idx + 2 < len(rows) and header_idx - banner_from < 4:
        nxt = rows[header_idx + 1]
        if not (_nonempty_count(header_row) == 1 and _nonempty_count(nxt) > 1 and _row_all_text(nxt)):
            break
        header_idx += 1
        header_row = rows[header_idx]
    if header_idx > banner_from:
        issues.append(
            f"row(s) {banner_from + 1}-{header_idx} look like title/banner rows; "
            f"using row {header_idx + 1} as the header row"
        )

    nxt = rows[header_idx + 1] if header_idx + 1 < len(rows) else None
    if nxt is not None and _row_all_text(nxt) and len(rows) > header_idx + 2:
        hr_n, nxt_n = _nonempty_count(header_row), _nonempty_count(nxt)
        if hr_n < nxt_n and hr_n > 1:
            # two-tier header: forward-fill the top tier and combine
            filled: list[object] = []
            last = None
            for v in header_row:
                if not _is_empty(v):
                    last = v
                filled.append(last)
            header_row = [
                f"{_cell_text(a)} {_cell_text(b)}".strip() for a, b in zip(filled, nxt, strict=False)
            ]
            issues.append(
                f"multi-row header detected: rows {header_idx + 1}-{header_idx + 2} combined"
            )
            header_idx += 1
            combined = True

    headers = header_row if combined else rows[header_idx]
    header_texts = [_cell_text(v) for v in headers]
    body = rows[header_idx + 1:]
    columns = _build_columns(header_texts, body)
    sample_rows = [_stringify_row(r) for r in body[:MAX_SAMPLE_ROWS]]
    _detect_subtotal_rows(body, header_idx + 1, issues)

    roles = [c.role for c in columns]
    has_member = ColumnRole.member in roles
    has_parent = ColumnRole.parent in roles
    level_count = sum(1 for r in roles if r == ColumnRole.level)

    # 1) chart of accounts: member+parent columns, or Level-1..Level-N columns
    if (has_member and has_parent) or level_count >= 2:
        hierarchy = coa.build_hierarchy(columns, body, sheet_name=sheet_name, header_rows=header_idx + 1)
        return SheetAnalysis(
            name=sheet_name, kind=SheetKind.chart_of_accounts, columns=columns,
            hierarchy=hierarchy, formulas=formulas, sample_rows=sample_rows, issues=issues,
        )

    # 2) layout / data table: a row of period-like column headers
    period_idx = _find_period_row(rows)
    if period_idx is not None:
        p_headers = [_cell_text(v) for v in rows[period_idx]]
        p_body = rows[period_idx + 1:]
        p_columns = _build_columns(p_headers, p_body)
        period_cols = [c for c in p_columns if c.role == ColumnRole.period]
        # De-duplicate: a daily sheet repeats "Actual" once per column, which
        # would otherwise fill the hint list with 260 copies of one word.
        seen_hints: set[str] = set()
        pov_hints: list[str] = []
        for r in rows[:period_idx]:
            for v in r:
                if _is_empty(v):
                    continue
                text = _cell_text(v)
                if text.lower() not in seen_hints:
                    seen_hints.add(text.lower())
                    pov_hints.append(text)
        pov_hints = pov_hints[:10]

        first_col = [r[0] for r in p_body]
        first_texts = [v for v in first_col if not _is_empty(v)]
        first_col_texty = bool(first_texts) and (
            sum(1 for v in first_texts if not _is_numericish(v)) / len(first_texts) >= 0.6
        )
        period_indexes = [c.index for c in period_cols]
        data_cells = [
            r[i] for r in p_body for i in period_indexes if i < len(r) and not _is_empty(r[i])
        ]
        numeric_frac = (
            sum(1 for v in data_cells if _is_numericish(v)) / len(data_cells) if data_cells else 0.0
        )

        if _is_empty(rows[period_idx][0]) and first_col_texty and p_body:
            layout = LayoutParse(
                row_labels=[_cell_text(v) for v in first_texts][:200],
                column_labels=[c.header for c in period_cols],
                pov_hints=pov_hints,
                issues=[] if numeric_frac >= 0.6 or not data_cells else
                ["grid body is not predominantly numeric; verify this really is a form layout"],
            )
            return SheetAnalysis(
                name=sheet_name, kind=SheetKind.layout, columns=p_columns, layout=layout,
                formulas=formulas, sample_rows=[_stringify_row(r) for r in p_body[:MAX_SAMPLE_ROWS]],
                issues=issues,
            )
        if numeric_frac >= 0.6 and p_body:
            label_col = next((c for c in p_columns if c.role == ColumnRole.label), None)
            label_idx = label_col.index if label_col else 0
            table = DataTableParse(
                period_columns=[c.header for c in period_cols],
                label_column=label_col.header if label_col else (p_columns[0].header or None),
                row_labels=[
                    _cell_text(r[label_idx]) for r in p_body
                    if label_idx < len(r) and not _is_empty(r[label_idx])
                    and not _is_numericish(r[label_idx])
                ][:200],
                row_count=sum(1 for r in p_body if _nonempty_count(r) > 0),
                sample_rows=[_stringify_row(r) for r in p_body[:MAX_SAMPLE_ROWS]],
                issues=[],
            )
            return SheetAnalysis(
                name=sheet_name, kind=SheetKind.data_table, columns=p_columns, data_table=table,
                formulas=formulas, sample_rows=table.sample_rows, issues=issues,
            )
        issues.append(
            "period-like column headers found but the body is neither a label grid nor mostly numeric; not classified"
        )

    if not any(i.startswith("period-like") for i in issues):
        issues.append(
            "could not classify sheet: no member+parent or level columns, and no period-like column headers"
        )
    return SheetAnalysis(
        name=sheet_name, kind=SheetKind.unknown, columns=columns,
        formulas=formulas, sample_rows=sample_rows, issues=issues,
    )


# --- workbook / file entry points -------------------------------------------


def _formula_text(value: object) -> str:
    if isinstance(value, str):
        return value
    text = getattr(value, "text", None)  # openpyxl ArrayFormula
    return text if isinstance(text, str) else str(value)


def _analyze_xlsx(path: Path, filename: str) -> WorkbookAnalysis:
    from openpyxl import load_workbook

    # data_only=False: formulas stay inert text and are never evaluated.
    wb = load_workbook(path, data_only=False, read_only=False, keep_vba=False, keep_links=False)
    try:
        sheets: list[SheetAnalysis] = []
        for ws in wb.worksheets:
            issues: list[str] = []
            formulas: list[FormulaInfo] = []
            merged = list(ws.merged_cells.ranges) if hasattr(ws, "merged_cells") else []
            if merged:
                sample = ", ".join(str(r) for r in merged[:5])
                issues.append(
                    f"{len(merged)} merged cell range(s) detected ({sample}); merged values live in the top-left cell only"
                )
            if ws.max_row and ws.max_row > MAX_ROWS:
                issues.append(f"sheet truncated to the first {MAX_ROWS} rows (has {ws.max_row})")
            if ws.max_column and ws.max_column > MAX_COLS:
                issues.append(f"sheet truncated to the first {MAX_COLS} columns (has {ws.max_column})")

            grid: list[list[object]] = []
            formula_cap_hit = False
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, MAX_ROWS),
                                    max_col=min(ws.max_column or 1, MAX_COLS)):
                values: list[object] = []
                for cell in row:
                    value = cell.value
                    if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                        text = _formula_text(value)
                        if len(formulas) < MAX_FORMULAS_PER_SHEET:
                            formulas.append(FormulaInfo(sheet=ws.title, cell=cell.coordinate, formula=text))
                        else:
                            formula_cap_hit = True
                        values.append(text if text.startswith("=") else f"={text}")
                    else:
                        values.append(value)
                grid.append(values)
            if formula_cap_hit:
                issues.append(f"formula extraction capped at {MAX_FORMULAS_PER_SHEET} formulas")
            sheets.append(_analyze_rows(ws.title, grid, formulas, issues))
    finally:
        wb.close()

    vba_modules: list[VbaModule] = []
    if path.suffix.lower() == ".xlsm":
        vba_modules, vba_issues = extract_vba_modules(path)
        if vba_issues and sheets:
            sheets[0].issues.extend(vba_issues)
    return WorkbookAnalysis(
        filename=filename, sheets=sheets, vba_modules=vba_modules,
        kind_guess=_workbook_kind(sheets),
    )


def _analyze_csv(path: Path, filename: str) -> WorkbookAnalysis:
    issues: list[str] = []
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(raw), dialect)
    rows: list[list[object]] = []
    for i, row in enumerate(reader):
        if i >= MAX_ROWS:
            issues.append(f"file truncated to the first {MAX_ROWS} rows")
            break
        rows.append([c if c.strip() else None for c in row[:MAX_COLS]])
    sheet = _analyze_rows(Path(filename).stem or "Sheet1", rows, [], issues)
    return WorkbookAnalysis(filename=filename, sheets=[sheet], kind_guess=_workbook_kind([sheet]))


def _workbook_kind(sheets: list[SheetAnalysis]) -> SheetKind:
    for kind in (SheetKind.chart_of_accounts, SheetKind.layout, SheetKind.data_table):
        if any(s.kind == kind for s in sheets):
            return kind
    return SheetKind.unknown


def analyze_file(path: Path, filename: str | None = None) -> WorkbookAnalysis:
    """Analyze a spreadsheet file deterministically. Parse only — never executes."""
    filename = filename or path.name
    ext = Path(filename).suffix.lower() or path.suffix.lower()
    if ext == ".csv":
        return _analyze_csv(path, filename)
    return _analyze_xlsx(path, filename)


# --- VBA (extraction only, never execution) ---------------------------------


def extract_vba_modules(path: Path, cap_chars: int = MAX_VBA_CHARS) -> tuple[list[VbaModule], list[str]]:
    """Extract VBA module source as inert, redacted text via oletools.

    The code is never compiled, interpreted or executed — it is surfaced so a
    human (or the chat skill) can read what the workbook would have run.
    """
    issues: list[str] = []
    modules: list[VbaModule] = []
    try:
        from oletools import olevba
    except ImportError:
        return [], ["oletools is not installed; VBA modules were not extracted"]
    try:
        parser = olevba.VBA_Parser(str(path))
        try:
            if parser.detect_vba_macros():
                total = 0
                for _fname, _stream, vba_filename, code in parser.extract_macros():
                    code = code or ""
                    remaining = cap_chars - total
                    if remaining <= 0:
                        issues.append("VBA extraction size cap reached; remaining modules skipped")
                        break
                    if len(code) > remaining:
                        code = code[:remaining]
                        issues.append(f"VBA module '{vba_filename}' truncated at the size cap")
                    total += len(code)
                    modules.append(VbaModule(
                        name=vba_filename or "Module",
                        line_count=code.count("\n") + 1 if code else 0,
                        code=redact_text(code),
                    ))
        finally:
            parser.close()
    except Exception as exc:  # noqa: BLE001 — a hostile/corrupt vbaProject must never break analysis
        issues.append(f"VBA extraction failed: {redact_text(str(exc))}")
    modules.sort(key=lambda m: m.name)
    return modules, issues
